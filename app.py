# app.py
import math, uuid, io, os, logging
from datetime import datetime
import tempfile, subprocess, shutil
from openpyxl.workbook.properties import CalcProperties
from openpyxl.cell.cell import MergedCell
import dash
from dash import html, dcc, Input, Output, State, dash_table, no_update
import dash_bootstrap_components as dbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
# ---- Logging (must be early) ----
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("doubleoak")

# ---- Configs ----
SIDEBAR_W = 360  # keep in sync with Offcanvas width (px)

# --- Pricebook loader: prefer core.pricebook if present, else fallback ---
try:
    from core import pricebook as pricebook  # type: ignore
except Exception:
    class _Pricebook:
        def __init__(self):
            self._path = os.environ.get("PRICEBOOK_PATH")
            self._loaded = False
            self._sheets = {}

        def ensure_loaded(self):
            if self._loaded:
                return
            self._sheets = {}
            p = self._path
            if p and os.path.exists(p):
                try:
                    self._sheets = pd.read_excel(p, sheet_name=None, engine="openpyxl")
                except Exception:
                    self._sheets = {}
            self._loaded = True

        def get_source(self):
            return self._path or "(defaults)"

        def get_price(self, sku, default):
            """Scan sheets for a row matching `sku` and return a price-like column; fallback to `default`."""
            self.ensure_loaded()
            key = str(sku or "").strip().lower()
            for _, df in (self._sheets or {}).items():
                try:
                    df2 = df.copy()
                    df2.columns = [str(c).strip().lower() for c in df2.columns]
                    price_cols = [c for c in df2.columns if c in (
                        "price","unit price","cost","rate","per lf","$/lf","$ / lf","$ per lf")]
                    key_cols = [c for c in df2.columns if c in (
                        "sku","key","code","item","name","description")]
                    for kc in key_cols:
                        series = df2[kc].astype(str).str.strip().str.lower()
                        mask = series == key
                        if mask.any() and price_cols:
                            val = df2.loc[mask, price_cols[0]].iloc[0]
                            return float(val)
                except Exception:
                    continue
            return default

    pricebook = _Pricebook()
# --- end pricebook fallback ---

def _write_cell(ws, coord: str, value, number_format: str | None = None):
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                cell = ws.cell(rng.min_row, rng.min_col)
                break
    cell.value = value
    if number_format is not None:
        cell.number_format = number_format

# ---- Proposal template bounds (rows for line items) ----
# Your template uses rows 14..40 for items; totals start at row 41 (J41..J44)
ITEM_START_ROW = int(os.environ.get("ITEM_START_ROW", "14"))
ITEM_END_ROW   = int(os.environ.get("ITEM_END_ROW",   "40"))

def populate_workbook(
    wb,
    lines,
    project_name,
    cat,
    sf_remove_tax,
    orange_remove_tax,
    *,
    hide_unused_rows: bool = False,   # ← only PDF will set this True
):
    ws = wb[wb.sheetnames[0]]

    # Title (B11)
    title = (project_name or "").strip() or "Untitled Project"
    existing = ws["B11"].value or "Project: "
    _write_cell(ws, "B11", f"Project: {title}" if str(existing).strip().lower().startswith("project:") else title)

    # ---- filter out "empty" rows before writing
    def _is_nonempty(row):
        if not row: return False
        item = str(row.get("Item") or "").strip()
        qty  = row.get("Qty")
        line = float(row.get("Line Total") or 0.0)
        # keep if: has an item name AND (qty > 0 OR line total > 0)
        if item and ((qty is not None and float(qty) > 0) or line > 0):
            return True
        return False

    clean_lines = [r for r in (lines or []) if _is_nonempty(r)]

    # ---- clear the item window ONLY (14..40) to avoid touching totals/shapes
    money_fmt = "$#,##0.00"
    for r in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        for c in ("C","D","H","I","J"):
            cell = ws[f"{c}{r}"]
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
        # also unhide in case it was hidden on a previous export
        ws.row_dimensions[r].hidden = False

    # ---- write rows up to capacity
    capacity = ITEM_END_ROW - ITEM_START_ROW + 1  # (40-14+1) = 27 rows
    used = min(len(clean_lines), capacity)
    if len(clean_lines) > capacity:
        # don’t overflow into totals – just log
        logger.warning("populate_workbook: %s items truncated to capacity %s (rows %s-%s)",
                       len(clean_lines), capacity, ITEM_START_ROW, ITEM_END_ROW)

    for i in range(used):
        r = ITEM_START_ROW + i
        row = clean_lines[i]
        _write_cell(ws, f"C{r}", row.get("Qty"))
        _write_cell(ws, f"D{r}", row.get("Item"))
        _write_cell(ws, f"H{r}", row.get("Unit"))
        _write_cell(ws, f"I{r}", float(row.get("Price Each") or 0.0), money_fmt)
        _write_cell(ws, f"J{r}", float(row.get("Line Total") or 0.0), money_fmt)

    # ---- optionally hide unused rows (PDF wants this)
    if hide_unused_rows and used < capacity:
        for r in range(ITEM_START_ROW + used, ITEM_END_ROW + 1):
            ws.row_dimensions[r].hidden = True
            # extra belt & suspenders for some PDF engines:
            # ws.row_dimensions[r].height = 0.1

    # ---- totals J41–J44 (unchanged)
    grand_subtotal = sum(float(x.get("Line Total") or 0.0) for x in clean_lines[:used])
    if (cat or "") == "Silt Fence":
        remove_tax_flag = "remove_tax" in (sf_remove_tax or [])
    elif (cat or "") == "Plastic Orange Fence":
        remove_tax_flag = "remove_tax" in (orange_remove_tax or [])
    else:
        remove_tax_flag = False
    tax_rate = 0.0 if remove_tax_flag else SALES_TAX
    sales_tax_total = grand_subtotal * tax_rate
    grand_total = grand_subtotal + sales_tax_total

    _write_cell(ws, "J41", float(grand_subtotal), money_fmt)
    _write_cell(ws, "J42", float(tax_rate), "0.00%")
    j43 = ws["J43"].value
    if not (isinstance(j43, str) and str(j43).startswith("=")):
        _write_cell(ws, "J43", float(sales_tax_total), money_fmt)
    else:
        ws["J43"].number_format = money_fmt
    j44 = ws["J44"].value
    if not (isinstance(j44, str) and str(j44).startswith("=")):
        _write_cell(ws, "J44", float(grand_total), money_fmt)
    else:
        ws["J44"].number_format = money_fmt

    return title.replace(" ", "_")

# --- Version helper (safe during import) ---
def _read_version_fallback() -> str:
    v = os.environ.get("APP_VERSION")
    if v:
        return v.strip()
    try:
        with open("VERSION", "r", encoding="utf-8") as f:
            return f.read().strip() or "0.0.0"
    except Exception:
        return "0.0.0"

# ---- Load pricebook at startup (root Excel or env path) ----
pricebook.ensure_loaded()
PRICEBOOK_SOURCE = pricebook.get_source()

# ---- Proposal template resolution + startup logging ----
DEFAULT_PROPOSAL_TEMPLATE = r"Z:\Double Oak Erosion\BIDS\proposal_template.xlsx"

def _resolve_proposal_template_path():
    candidates = [
        ("env:PROPOSAL_TEMPLATE_PATH", (os.environ.get("PROPOSAL_TEMPLATE_PATH", "").strip() or None)),
        ("hardcoded Windows Z:", DEFAULT_PROPOSAL_TEMPLATE),
        ("fallback /mnt/data", "/mnt/data/proposal_template.xlsx"),
    ]
    checked = []
    chosen = None
    for label, p in candidates:
        if p:
            exists = os.path.exists(p)
            checked.append(f"{label} -> {p} (exists={exists})")
            if exists and chosen is None:
                chosen = p
        else:
            checked.append(f"{label} -> (unset)")
    return chosen, checked

def _log_startup_paths():
    pricebook_path = os.environ.get("PRICEBOOK_PATH", "").strip() or "(unset)"
    logger.info(f"PRICEBOOK_PATH = {pricebook_path} (exists={os.path.exists(pricebook_path) if pricebook_path!='(unset)' else False})")
    chosen, checked = _resolve_proposal_template_path()
    logger.info("Proposal template path resolution:")
    for line in checked:
        logger.info("  " + line)
    logger.info(f"Using proposal template: {chosen if chosen else '(not found)'}")

try:
    _log_startup_paths()
except Exception as e:
    logger.warning(f"Startup path logging failed: {e}")

# ---- Constants / SKUs ----
FABRIC_SKU_14G = "silt-fence-14g"
FABRIC_SKU_125G = "silt-fence-12g5"
FABRIC_SKU_UNREINF = "silt-fence-unreinforced"        # <-- new
POST_SKU_T_POST_4FT = "t-post-4ft"
POST_SKU_TXDOT_T_POST_4FT = "tx-dot-t-post-4-ft"
POST_SKU_T_POST_6FT = "t-post-6ft"
POST_SKU_WOOD_STAKE_4FT = "wood-stake-4ft"            # <-- new
FABRIC_SKU_ORANGE_LIGHT = "orange-fence-light-duty"
FABRIC_SKU_ORANGE_HEAVY = "orange-fence-heavy-duty"
CAP_SKU_OSHA = "cap-osha"
CAP_SKU_PLASTIC = "cap-plastic"

SALES_TAX = 0.0825
PROD_LF_PER_DAY = 2500

def required_footage(total_lf: float, waste_pct: float) -> float:
    return max(0.0, float(total_lf or 0)) * (1.0 + max(0.0, float(waste_pct or 0))/100.0)

def posts_needed(required_ft: float, spacing_ft: int) -> int:
    rf=max(0.0, float(required_ft or 0)); sp=max(1, int(spacing_ft or 1))
    return int(math.ceil(rf/sp)) + (1 if rf>0 else 0)

def rolls_needed(required_ft: float, roll_len: int=100) -> int:
    rf=max(0.0,float(required_ft or 0)); rl=max(1,int(roll_len or 100))
    return int(math.ceil(rf/rl)) if rf>0 else 0

def get_labor_per_day()->float: return 554.34
def fuel_cost(days:int, any_work:bool)->float: return (65.0*max(0,int(days or 0))) if any_work else 0.0

def materials_breakdown(required_ft: float, cost_per_lf: float, posts_count: int, post_unit_cost: float, tax_rate: float=SALES_TAX):
    fabric_cost = max(0.0,required_ft)*max(0.0,cost_per_lf)
    hardware_cost = max(0,posts_count)*max(0.0,post_unit_cost)
    subtotal = fabric_cost + hardware_cost
    tax = subtotal*tax_rate
    return fabric_cost, hardware_cost, subtotal, tax

# ---- Theme (Bootstrap + CSS variables) ----
external_stylesheets = [dbc.themes.BOOTSTRAP]
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
app.title = f"Double Oak Fencing Estimator (Dash) – v{_read_version_fallback()}"

app.index_string = """
<!DOCTYPE html>
<html>
<head>
  {%metas%}
  <title>{%title%}</title>
  {%favicon%}
  {%css%}
  <style>
    .theme-light {
      --bg: #C3D4B3; --card: #E0E9D8; --text: #0B1F18; --accent: #1F8A3B; --muted: #CFE9D3;
      --header-bg: #2E7D32; --header-text: #1C2515; --table-stripe: #EFF7F0; --badge-grad: linear-gradient(90deg, #2E7D32, #145A32);
      --label-color: #2E7D32;
    }
    .theme-dark {
      --bg: #112813; --card: #132118; --text: #ffffff; --accent: #65824a; --muted: #020603;
      --header-bg: #14532D; --header-text: #E6F4EA; --table-stripe: #020406; --badge-grad: linear-gradient(90deg, #059669, #065F46);
      --label-color: #22C55E;
    }
    html, body { background: transparent !important; }
    #theme_root { background: var(--bg); color: var(--text); min-height: 100vh; }
    .do-card { border: 2px solid var(--accent); border-radius: 12px; background: var(--card); box-shadow: 0 2px 10px rgba(0,0,0,.06); }
    .do-title { font-weight: 800; color: var(--text); border-bottom: 2px dashed var(--muted); padding-bottom: 6px; margin-bottom: 12px; }
    .status-badge { border: 2px solid var(--accent); border-radius: 10px; padding: 10px 24px; font-weight: 800; display: inline-block; background: var(--badge-grad); color: #fff; }
    .offcanvas.theme-surface { background: var(--card) !important; border-right: 3px solid var(--accent); }
    #theme_root .offcanvas.theme-surface .offcanvas-body label,
    #theme_root .offcanvas.theme-surface .offcanvas-body .form-label,
    #theme_root .offcanvas.theme-surface .offcanvas-body .form-check-label { color: var(--label-color) !important; }
    #theme_root .offcanvas.theme-surface .offcanvas-title { color: var(--label-color) !important; }
    #theme_root .offcanvas.theme-surface .offcanvas-header { color: var(--header-text) !important; border-bottom: 1px solid var(--muted); }
    #theme_root .offcanvas.theme-surface .offcanvas-title { color: var(--header-text) !important; }
    .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner th {
      background: var(--header-bg) !important; color: var(--header-text) !important; font-weight: 700;
    }
    .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner td { color: var(--text); }
    .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner tr:nth-child(odd) td { background: var(--table-stripe); }
    .btn-success { background-color: var(--accent); border-color: var(--accent); }
    .btn-success:hover { filter: brightness(0.95); }
    .floating-btn { position: fixed; z-index: 20000; box-shadow: 0 2px 8px rgba(0,0,0,.15); }
  </style>
</head>
<body>
  {%app_entry%}
  <footer>{%config%}{%scripts%}{%renderer%}</footer>
</body>
</html>
"""

# ---- Sidebar controls ----
sidebar = dbc.Offcanvas(
    id="sidebar",
    title="Fencing Estimator",
    is_open=True,
    placement="start",
    scrollable=True,
    close_button=False,
    backdrop=False,
    keyboard=False,
    class_name="theme-surface",
    style={"width": f"{SIDEBAR_W}px"},
    children=[
        html.Div(
            [
                html.Div("Profit", style={"fontWeight": 700, "fontSize": "12px", "lineHeight": "1.0", "marginBottom": "2px"}),
                html.Div(id="profit_status", style={"marginTop": "2px"}),
                html.Div(id="profit_pill_css", style={"marginTop": "2px"})
            ],
            style={"marginBottom": "8px", "lineHeight": "1.1"}
        ),
        html.Hr(),
        dbc.Label("Project Title"),
        dcc.Input(id="project_name", type="text", placeholder="Lakeside Retail – Phase 2", style={"width":"100%"}),
        # (customer/address intentionally omitted)
        html.Hr(),
        dbc.Label("Fencing Material"),
        dcc.Dropdown(
            id="fence_category",
            options=[{"label":"Silt Fence","value":"Silt Fence"}, {"label":"Plastic Orange Fence","value":"Plastic Orange Fence"}],
            value="Silt Fence", clearable=False
        ),
        dbc.Label("Total Job Footage (ft)"),
        dcc.Input(id="total_lf", type="number", min=0, step=1, value=1000, style={"width":"100%"}),
        dbc.Label("Waste %"),
        dcc.Slider(id="waste_pct", min=0, max=10, step=0.5, value=2, marks=None, tooltip={"placement":"bottom","always_visible":True}),
        html.Div(id="silt_opts", children=[
            dbc.Label("Silt Fence Gauge"),
            dcc.Dropdown(
                id="sf_gauge",
                options=[
                    {"label":"14 Gauge (Reinforced)","value":"14 Gauge"},
                    {"label":"12.5 Gauge (Reinforced)","value":"12.5 Gauge"},
                    {"label":"Unreinforced (Wood Stakes)","value":"Unreinforced"},  # <-- new
                ],
                value="14 Gauge", clearable=False
            ),
            dbc.Label("Post/Stake Spacing (ft)"),
            dcc.Dropdown(id="sf_spacing", options=[3,4,6,8,10], value=8, clearable=False),
            dbc.Label("Final Price / LF"),
            dcc.Input(id="sf_price_lf", type="number", min=0, step=0.01, value=2.50, style={"width":"100%"}),
            dbc.Checklist(options=[{"label": "Add safety caps","value":"caps"}], value=[], id="sf_caps"),
            html.Div(
                id="sf_cap_wrap",
                style={"display": "none"},
                children=[
                    dbc.Label("Cap Type"),
                    dcc.Dropdown(
                        id="sf_cap_type",
                        options=[{"label":"OSHA-Approved","value":"OSHA"}, {"label":"Regular Plastic Cap","value":"PLASTIC"}],
                        value="OSHA", clearable=False
                    ),
                ],
            ),
            dbc.Checklist(options=[{"label":"Add fence removal pricing","value":"removal"}], value=[], id="sf_removal"),
            dbc.Checklist(options=[{"label":"Remove sales tax from customer printout","value":"remove_tax"}], value=[], id="sf_remove_tax"),
        ]),
        html.Div(id="orange_opts", children=[
            dbc.Label("Orange Fence Duty"),
            dcc.Dropdown(id="orange_duty", options=[{"label":"Light Duty","value":"Light Duty"},{"label":"Heavy Duty","value":"Heavy Duty"}], value="Light Duty", clearable=False),
            dbc.Label("T-Post Spacing (ft)"),
            dcc.Dropdown(id="orange_spacing", options=[3,4,6,8,10], value=10, clearable=False),
            dbc.Label("Final Price / LF"),
            dcc.Input(id="orange_price_lf", type="number", min=0, step=0.01, value=2.50, style={"width":"100%"}),
            dbc.Checklist(options=[{"label":"Add fence removal pricing","value":"removal"}], value=[], id="orange_removal"),
            dbc.Checklist(options=[{"label":"Remove sales tax from customer printout","value":"remove_tax"}], value=[], id="orange_remove_tax"),
        ])
    ],
)

# ---- Cards ----
cards = dbc.Row([
    dbc.Col(dbc.Card([dbc.CardBody([html.H4("Cost Summary", className="do-title"), html.Div(id="cost_summary")])], className="do-card"), md=4),
    dbc.Col(dbc.Card([dbc.CardBody([html.H4("Total Costs Breakdown", className="do-title"), html.Div(id="total_costs")])], className="do-card"), md=4),
    dbc.Col(dbc.Card([dbc.CardBody([html.H4("Material Cost Breakdown", className="do-title"), html.Div(id="material_costs")])], className="do-card"), md=4),
], className="g-4")

# ---- Table + Export ----
table_section = dbc.Row([
    dbc.Col([
        html.Div(
            [
                dbc.Button("⬇️ Export Proposal (Excel)", id="btn_download_xlsx", color="success", class_name="mb-2"),
                dbc.Button("🖨️ Export Proposal (PDF)", id="btn_download_pdf", color="secondary", class_name="mb-2 ms-2"),
            ],
            style={"display":"flex","justifyContent":"flex-end"}
        ),
        html.H4("📑 Customer Printout Preview", className="do-title"),
        dash_table.DataTable(
            id="preview_table",
            columns=[
                {"name":"Qty","id":"Qty","type":"numeric"},
                {"name":"Item","id":"Item"},
                {"name":"Unit","id":"Unit"},
                {"name":"Price Each","id":"Price Each","type":"numeric","format":dash_table.FormatTemplate.money(2)},
                {"name":"Line Total","id":"Line Total","type":"numeric","format":dash_table.FormatTemplate.money(2)},
            ],
            data=[], editable=False, row_deletable=True,
            style_table={"overflowX":"auto"},
            style_cell={"fontFamily":"Inter, system-ui, -apple-system, Segoe UI, Roboto","fontSize":"16px","padding":"8px"},
            style_header={"fontWeight":"700"},
            style_data_conditional=[]
        ),
        html.Div(id="totals_right", style={"textAlign":"right","marginTop":"10px","fontWeight":"700"})
    ], md=12)
])

# -- helper to position the hamburger tab on/off canvas edge
def _tab_style(is_open: bool):
    return {
        "position": "fixed",
        "top": "329px",
        "left": f"{SIDEBAR_W - 15}px" if is_open else "0px",
        "zIndex": 20000,
        "boxShadow": "0 2px 8px rgba(0,0,0,.15)",
        "borderTopRightRadius": "50px",
        "borderBottomRightRadius": "50px",
        "borderTopLeftRadius": "10px",
        "borderBottomLeftRadius": "10px",
        "padding": "8px 12px",
    }

# ---- Main wrapper ----
main_wrap = html.Div(
    [
        html.Br(),
        cards,
        html.Br(),
        html.Hr(),
        table_section,
        html.Footer(
    dbc.Badge(
        [
            html.Span("✅", style={"marginRight": "8px"}),  # green check only here
            html.Span(f"Double Oak Estimator – v{_read_version_fallback()}"),
        ],
        color="secondary",
        pill=True,
        class_name="shadow-sm",
        style={"display": "inline-flex", "alignItems": "center"}
    ),
    style={"position": "fixed", "bottom": "10px", "right": "12px", "zIndex": 9999, "background": "transparent"}
)
    ],
    id="main_wrap",
    style={"marginLeft": f"{SIDEBAR_W}px", "transition": "margin-left .25s ease"}
)

# ---- THEME ROOT ----
theme_root = html.Div(
    [
        html.Button("🌙", id="theme_toggle", n_clicks=0,
                    className="btn btn-outline-secondary floating-btn",
                    style={"top":"12px","right":"12px"}),

        html.Button("☰", id="open_sidebar_btn", n_clicks=0,
                    className="btn btn-success floating-btn",
                    style=_tab_style(True)),
        sidebar,
        main_wrap,

        # Toast container for download messages
        html.Div(id="toast_container"),
    ],
    id="theme_root",
    className="theme-light"
)

# ---- Layout ----
app.layout = dbc.Container(
    [
        dcc.Store(id="theme_store", data="light", storage_type="local"),
        theme_root,
        dcc.Download(id="dl_proposal_xlsx"),  # download target
        dcc.Download(id="dl_proposal_pdf"),
    ],
    fluid=True, className="p-0", style={"maxWidth": "100%", "padding": 0}
)

def tiny_profit_pill(value_pct: float, *, target_pct: float = 30.0, width_px: int = 100, height_px: int = 10):
    value_pct = max(0.0, min(100.0, float(value_pct or 0.0)))
    track_style = {"height": f"{height_px}px","borderRadius": "9999px",
                   "background": "linear-gradient(90deg, #ef4444 0%, #f97316 20%, #f59e0b 35%, #facc15 50%, #a3e635 70%, #22c55e 100%)","opacity": "0.95"}
    target_style = {"position":"absolute","top": f"{-(height_px//4)}px","left": f"{target_pct:.2f}%",
                    "height": f"{height_px + (height_px//2)}px","borderLeft": "2px dashed #ff9d00"}
    needle_style = {"position": "absolute","top": f"{-(height_px//4)}px","left": f"{value_pct:.2f}%",
                    "height": f"{height_px + (height_px//2)}px","width":"2px","background":"#0f172a","borderRadius":"1px"}
    wrap_style = {"position":"relative","width": f"{width_px}px","height": f"{height_px}px","margin":"4px auto","overflow":"hidden"}
    return html.Div(style=wrap_style, children=[html.Div(style=track_style), html.Div(style=target_style), html.Div(style=needle_style)])

# ---- Callbacks ----
@app.callback(
    Output("sf_cap_wrap", "style"),
    Input("sf_caps", "value"),
    State("sf_gauge", "value")
)
def toggle_cap_type(caps_values, gauge):
    show = ("caps" in (caps_values or [])) and (gauge != "Unreinforced")
    return {} if show else {"display": "none"}

@app.callback(
    Output("silt_opts","style"),
    Output("orange_opts","style"),
    Input("fence_category","value")
)
def toggle_category(cat):
    if cat == "Silt Fence":
        return ({}, {"display":"none"})
    return ({"display":"none"}, {})

@app.callback(
    Output("cost_summary","children"),
    Output("total_costs","children"),
    Output("material_costs","children"),
    Output("profit_status","children"),
    Output("profit_pill_css","children"),
    Output("preview_table","data"),
    Output("totals_right","children"),
    Input("fence_category","value"),
    Input("total_lf","value"),
    Input("waste_pct","value"),
    Input("sf_gauge","value"),
    Input("sf_spacing","value"),
    Input("sf_price_lf","value"),
    Input("sf_caps","value"),
    Input("sf_cap_type","value"),
    Input("sf_removal","value"),
    Input("sf_remove_tax","value"),
    Input("orange_duty","value"),
    Input("orange_spacing","value"),
    Input("orange_price_lf","value"),
    Input("orange_removal","value"),
    Input("orange_remove_tax","value"),
)

def compute(cat, total_lf, waste_pct, sf_gauge, sf_spacing, sf_price_lf, sf_caps, sf_cap_type,
            sf_removal, sf_remove_tax, orange_duty, orange_spacing, orange_price_lf,
            orange_removal, orange_remove_tax):

    total_lf = int(total_lf or 0)
    waste_pct = int(waste_pct or 0)
    remove_tax_flag = False
    include_caps = False
    cap_type = None
    final_price_per_lf = 2.50

    if cat == "Silt Fence":
        post_spacing = int(sf_spacing or 8)
        final_price_per_lf = float(sf_price_lf or 2.50)
        # Unreinforced uses wood stakes; ignore caps
        include_caps = ("caps" in (sf_caps or [])) and (sf_gauge != "Unreinforced")
        cap_type = sf_cap_type
        remove_tax_flag = ("remove_tax" in (sf_remove_tax or []))
        if (sf_gauge or "") == "Unreinforced":
            fabric_sku, fabric_default = FABRIC_SKU_UNREINF, 0.28
            post_sku, post_default = POST_SKU_WOOD_STAKE_4FT, 1.25
        elif (sf_gauge or "").startswith("14"):
            fabric_sku, fabric_default = FABRIC_SKU_14G, 0.32
            post_sku, post_default = POST_SKU_T_POST_4FT, 1.80
        else:
            fabric_sku, fabric_default = FABRIC_SKU_125G, 0.38
            post_sku, post_default = POST_SKU_TXDOT_T_POST_4FT, 2.15
        removal_selected = ("removal" in (sf_removal or []))
    else:
        post_spacing = int(orange_spacing or 10)
        final_price_per_lf = float(orange_price_lf or 2.50)
        remove_tax_flag = ("remove_tax" in (orange_remove_tax or []))
        if (orange_duty or "").startswith("Light"):
            fabric_sku, fabric_default = FABRIC_SKU_ORANGE_LIGHT, 0.30
        else:
            fabric_sku, fabric_default = FABRIC_SKU_ORANGE_HEAVY, 0.45
        post_sku, post_default = POST_SKU_T_POST_6FT, 2.25
        removal_selected = ("removal" in (orange_removal or []))

    # --- prices
    cost_per_lf    = pricebook.get_price(fabric_sku,  fabric_default) or fabric_default
    post_unit_cost = pricebook.get_price(post_sku,    post_default)   or post_default
    caps_unit_cost = 0.0
    if cat=="Silt Fence" and include_caps and cap_type:
        caps_unit_cost = pricebook.get_price(CAP_SKU_OSHA if cap_type=="OSHA" else CAP_SKU_PLASTIC,
                                             3.90 if cap_type=="OSHA" else 1.05) or (3.90 if cap_type=="OSHA" else 1.05)

    # --- calcs
    req_ft = required_footage(total_lf, waste_pct)
    posts = posts_needed(req_ft, post_spacing)
    rolls = rolls_needed(req_ft)
    caps_qty = posts if (cat=="Silt Fence" and include_caps and cap_type) else 0
    caps_cost = caps_qty * caps_unit_cost

    fabric_cost, hardware_cost, mat_sub, tax = materials_breakdown(req_ft, cost_per_lf, posts, post_unit_cost, SALES_TAX)
    mat_sub_all = mat_sub + caps_cost
    tax_all = tax + caps_cost*SALES_TAX

    project_days = (req_ft/PROD_LF_PER_DAY) if req_ft>0 else 0.0
    labor_cost = project_days*get_labor_per_day()
    billing_days = math.ceil(project_days) if req_ft>0 else 0
    fuel = fuel_cost(billing_days, any_work=req_ft>0)

    def _calc_removal(req_ft: float, sell_per_lf: float):
        if req_ft <= 0:
            return 0.0, 0.0
        base = sell_per_lf * 0.40
        if req_ft < 800: floor = 1.15
        elif req_ft < 2000: floor = 0.90
        elif req_ft < 10000:
            slope = (0.80 - 0.90) / (10000 - 2000)
            floor = 0.90 + slope * (req_ft - 2000)
        else:
            floor = 0.80
        unit = max(base, floor)
        total = unit * req_ft
        if total < 800.0:
            total = 800.0
            unit = total / req_ft
        return unit, total

    removal_unit_lf, removal_total = _calc_removal(req_ft, final_price_per_lf) if removal_selected else (0.0, 0.0)

    sell_total_main = final_price_per_lf * req_ft if req_ft>0 else 0.0
    caps_revenue = caps_unit_cost * caps_qty if caps_qty>0 else 0.0
    removal_revenue = removal_total if (removal_selected and req_ft>0) else 0.0
    customer_subtotal_display = sell_total_main + caps_revenue + removal_revenue
    customer_sales_tax = 0.0 if remove_tax_flag else customer_subtotal_display * SALES_TAX
    customer_total = customer_subtotal_display + customer_sales_tax

    internal_total_cost = mat_sub_all + tax_all + labor_cost + fuel
    subtotal_for_margin = sell_total_main + caps_revenue
    gross_profit = subtotal_for_margin - internal_total_cost
    profit_margin = (gross_profit / subtotal_for_margin) if subtotal_for_margin>0 else 0.0

    # ---- Panels (HTML)
    cs = dbc.Table([
        html.Tbody([
            html.Tr([html.Td("Subtotal (excl. sales tax)"), html.Td(f"${customer_subtotal_display:,.2f}", style={"textAlign":"right"})]),
            html.Tr([html.Td(f"Sales Tax ({0 if remove_tax_flag else SALES_TAX*100:.2f}%)"), html.Td(f"${customer_sales_tax:,.2f}", style={"textAlign":"right"})]),
            html.Tr([html.Td(html.Strong("Customer Total")), html.Td(html.Strong(f"${customer_total:,.2f}"), style={"textAlign":"right"})]),
            html.Tr([html.Td("Gross Profit"), html.Td(f"${gross_profit:,.2f}", style={"textAlign":"right"})]),
        ])
    ], bordered=False, striped=True, hover=False, size="sm")

    tc_rows = [
        html.Tr([html.Td("Total Material Cost"), html.Td(f"${mat_sub_all:,.2f}", style={"textAlign":"right"})]),
        html.Tr([html.Td("Labor Cost"), html.Td(f"${labor_cost:,.2f}", style={"textAlign":"right"})]),
        html.Tr([html.Td("Fuel"), html.Td(f"${fuel:,.2f}", style={"textAlign":"right"})]),
    ]
    if removal_selected and req_ft>0:
        tc_rows.append(html.Tr([html.Td("Fence Removal"), html.Td(f"${removal_total:,.2f}", style={"textAlign":"right"})]))
    tc_rows.append(html.Tr([html.Td("Final Price / LF (sell)"), html.Td(f"${final_price_per_lf:,.2f}", style={"textAlign":"right"})]))
    tc = dbc.Table(html.Tbody(tc_rows), bordered=False, striped=True, hover=False, size="sm")

    mc_rows = [
        html.Tr([html.Td(("Fabric (Silt Fence)" if cat=="Silt Fence" else "Plastic Orange Fence")), html.Td(f"${fabric_cost:,.2f}", style={"textAlign":"right"})]),
        html.Tr([html.Td("Posts / Stakes"), html.Td(f"${hardware_cost:,.2f}", style={"textAlign":"right"})]),
    ]
    if caps_qty>0:
        mc_rows.append(html.Tr([html.Td("Safety Caps"), html.Td(f"${caps_cost:,.2f}", style={"textAlign":"right"})]))
    mc_rows.extend([
        html.Tr([html.Td("Total Material Cost"), html.Td(f"${mat_sub_all:,.2f}", style={"textAlign":"right"})]),
        html.Tr([html.Td("Total Material Cost / LF"), html.Td(f"${(mat_sub_all/req_ft) if req_ft>0 else 0.0:,.2f}", style={"textAlign":"right"})]),
    ])
    mc = dbc.Table(html.Tbody(mc_rows), bordered=False, striped=True, hover=False, size="sm")

    # ---- Profit badge + tiny pill
    badge = html.Span(
        f" {'GOOD' if profit_margin>=0.30 else 'CHECK PROFIT'}  {profit_margin*100:.1f}%",
        className="status-badge",
        style={"fontSize": "15px", "padding": "6px 10px"}
    )
    pill = tiny_profit_pill((profit_margin*100.0) if subtotal_for_margin>0 else 0.0, target_pct=30.0, width_px=150, height_px=20)

    # ---- Customer lines
    lines = []
    customer_qty = int(total_lf or 0)
    if customer_qty>0:
        if cat=="Silt Fence":
            if (sf_gauge or "") == "Unreinforced":
                item_name = "Unreinforced Silt Fence (Wood Stakes)"
            elif (sf_gauge or "").startswith("14"):
                item_name = "14 Gauge Silt Fence"
            else:
                item_name = "12.5 Gauge Silt Fence"
        else:
            item_name = "Plastic Orange Fence"
        lines.append({"_id":str(uuid.uuid4()), "Qty":customer_qty, "Item":item_name, "Unit":"LF",
                      "Price Each":float(final_price_per_lf), "Line Total":float(final_price_per_lf)*customer_qty})
    if caps_qty>0:
        lines.append({"_id":str(uuid.uuid4()), "Qty":int(caps_qty), "Item":"Safety Caps", "Unit":"EA",
                      "Price Each":float(caps_unit_cost), "Line Total":float(caps_unit_cost)*int(caps_qty)})
    if removal_selected and req_ft>0:
        lines.append({"_id":str(uuid.uuid4()), "Qty":customer_qty, "Item":"Fence Removal", "Unit":"LF",
                      "Price Each":float(removal_unit_lf), "Line Total":float(removal_unit_lf)*customer_qty})

    # ---- UI totals mirroring J41–J44
    grand_subtotal = sum(float(l["Line Total"]) for l in lines)
    tax_rate = 0.0 if remove_tax_flag else SALES_TAX
    sales_tax_total = grand_subtotal * tax_rate
    grand_total = grand_subtotal + sales_tax_total
    totals_html = html.Div([
        html.Div(f"Grand Subtotal: ${grand_subtotal:,.2f}"),
        html.Div(f"Tax Rate: {tax_rate*100:.2f}%{' (removed)' if remove_tax_flag else ''}"),
        html.Div(f"Sales Tax: ${sales_tax_total:,.2f}"),
        html.Div(html.Strong(f"Grand Total: ${grand_total:,.2f}")),
    ], style={"textAlign":"right"})

    return cs, tc, mc, badge, pill, lines, totals_html

# -- Toggle Offcanvas
@app.callback(Output("sidebar", "is_open"), Input("open_sidebar_btn", "n_clicks"), State("sidebar", "is_open"), prevent_initial_call=True)
def toggle_sidebar(n, is_open):
    if not n: return no_update
    return not bool(is_open)

# -- Reposition the hamburger tab
@app.callback(Output("open_sidebar_btn", "style"), Input("sidebar", "is_open"))
def position_tab(is_open): return _tab_style(bool(is_open))

# -- Shift main content when sidebar opens/closes
@app.callback(Output("main_wrap", "style"), Input("sidebar", "is_open"))
def shift_main(is_open): return {"marginLeft": f"{SIDEBAR_W}px" if is_open else "0px", "transition": "margin-left .25s ease"}

# -- Theme: A) toggle stored mode on click
@app.callback(Output("theme_store", "data"), Input("theme_toggle", "n_clicks"), State("theme_store", "data"), prevent_initial_call=True)
def switch_theme(n, mode): return "dark" if (mode or "light") == "light" else "light"

# -- Theme: B) apply mode to UI
@app.callback(Output("theme_root", "className"), Output("theme_toggle", "children"), Input("theme_store", "data"))
def apply_theme(mode): return ("theme-dark","☀️") if (mode or "light")=="dark" else ("theme-light","🌙")

# ---- Download Excel (proposal) ----
@app.callback(
    Output("dl_proposal_xlsx", "data"),
    Input("btn_download_xlsx", "n_clicks"),
    State("preview_table", "data"),
    State("project_name", "value"),
    State("fence_category", "value"),
    State("sf_remove_tax", "value"),
    State("orange_remove_tax", "value"),
    prevent_initial_call=True
)
def download_proposal(n, lines, project_name, cat, sf_remove_tax, orange_remove_tax):
    if not n:
        return no_update

    template_path, _ = _resolve_proposal_template_path()
    if not template_path:
        logger.error("download_proposal: template not found")
        return no_update

    try:
        wb = load_workbook(template_path, keep_links=False, data_only=False)
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception as e:
        logger.exception("download_proposal: failed to open template: %s", e)
        return no_update

    safe_name = populate_workbook(wb, lines, project_name, cat, sf_remove_tax, orange_remove_tax)

    def _write_wb(bio):
        wb.save(bio)

    logger.info("download_proposal: sending workbook to browser")
    return dcc.send_bytes(_write_wb, filename=f"Proposal_{safe_name}.xlsx")

@app.callback(
    Output("dl_proposal_pdf", "data"),
    Input("btn_download_pdf", "n_clicks"),
    State("preview_table", "data"),
    State("project_name", "value"),
    State("fence_category", "value"),
    State("sf_remove_tax", "value"),
    State("orange_remove_tax", "value"),
    prevent_initial_call=True
)
def download_proposal_pdf(n, lines, project_name, cat, sf_remove_tax, orange_remove_tax):
    if not n:
        return no_update
    if shutil.which("soffice") is None:
        logger.error("download_proposal_pdf: soffice not found in container")
        return no_update

    template_path, _ = _resolve_proposal_template_path()
    if not template_path:
        logger.error("download_proposal_pdf: template not found")
        return no_update

    try:
        wb = load_workbook(template_path, keep_links=False, data_only=False)
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception as e:
        logger.exception("download_proposal_pdf: failed to open template: %s", e)
        return no_update

    safe_name = populate_workbook(
    wb,
    lines,
    project_name,
    cat,
    sf_remove_tax,
    orange_remove_tax,
    hide_unused_rows=True,     # 👈 collapse blanks for PDF
)

    with tempfile.TemporaryDirectory() as td:
        xlsx_path = os.path.join(td, f"proposal_{uuid.uuid4().hex}.xlsx")
        wb.save(xlsx_path)
        cmd = ["soffice","--headless","--norestore","--nolockcheck","--convert-to","pdf","--outdir",td,xlsx_path]
        try:
            res = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120)
        except Exception as e:
            logger.exception("download_proposal_pdf: soffice run failed: %s", e)
            return no_update
        if res.returncode != 0:
            logger.error("download_proposal_pdf: soffice rc=%s, msg=%s",
                         res.returncode, (res.stderr or res.stdout).decode("utf-8","ignore")[:600])
            return no_update

        pdf_path = os.path.join(td, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
        if not os.path.exists(pdf_path):
            logger.error("download_proposal_pdf: expected pdf not produced")
            return no_update

        logger.info("download_proposal_pdf: sending pdf to browser")
        return dcc.send_file(pdf_path, filename=f"Proposal_{safe_name}.pdf")

# expose Flask server for gunicorn/hosted platforms
server = app.server

if __name__ == "__main__":
    app.run_server(host="0.0.0.0", port=8050, debug=True)
